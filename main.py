from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from dotenv import load_dotenv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import warnings, os
import time
from send_mail import send_mail

load_dotenv()
warnings.filterwarnings("ignore")


def website_login(driver):
    login_username = driver.find_element(By.ID, "j_username")
    login_username.send_keys(os.getenv("WEBSITE_USERNAME"))
    login_password = driver.find_element(By.ID, "j_password")
    login_password.send_keys(os.getenv("WEBSITE_PASSWORD"))
    login_submit = driver.find_element(By.ID, "logOnFormSubmit")
    login_submit.click()


def search1(driver):
    search_btn1 = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//button[@title="Search"]'))
    )
    search_btn1.click()

    search_field1 = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//input[@placeholder="Search"]'))
    )
    search_field1.send_keys("Vendor Rating Report")

    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located(
            (By.XPATH, '//div[@title="Vendor Rating Report"]')
        )
    )
    ActionChains(driver).send_keys(Keys.DOWN).send_keys(Keys.ENTER).perform()


def search2(driver):
    sap_home = driver.find_element(By.XPATH, '//img[@title="Home"]')
    sap_home.click()

    search_field2 = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//input[@placeholder="Search"]'))
    )
    search_field2.send_keys("PM Vendor SOB Report New")

    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located(
            (By.XPATH, '//div[@title="PM Vendor SOB Report New"]')
        )
    )
    ActionChains(driver).send_keys(Keys.DOWN, Keys.DOWN).send_keys(Keys.ENTER).perform()


def pass_story_credentials(driver):
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "__dialog0"))
    )
    story_username = driver.find_element(By.CLASS_NAME, "sapEpmUiInputTextArea")
    story_username.send_keys(os.getenv("STORY_USERNAME"))
    story_password = driver.find_element(By.XPATH, '//input[@type="password"]')
    story_password.send_keys(os.getenv("STORY_PASSWORD"))
    ok_btn = driver.find_element(By.XPATH, '//span[@class="sapMBtnContent"]')
    ok_btn.click()


def vendor_report_download(driver):
    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//html"))
    )
    plant_btn = driver.find_element(By.XPATH, '//span[@title="Plant(1)"]')
    plant_btn.click()
    all_option = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                '//div[@class="sapEPMUIEffectiveFilterWidgetItem sapEPMUIEffectiveFilterWidgetItem-allItem"]',
            )
        )
    )
    all_option.click()
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located((By.ID, "__indicator0-busy-area"))
    )
    plant_btn = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.ID, "__control4--TV"))
    )
    plant_btn.click()
    time.sleep(1)
    cal_year_month = driver.find_element(
        By.XPATH, '//div[@class="sapEPMUIFilterItem sapEPMUIFilterItem-selected"]'
    )
    driver.execute_script("arguments[0].click()", cal_year_month)
    all_option = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                '//div[@class="sapEPMUIEffectiveFilterWidgetItem sapEPMUIEffectiveFilterWidgetItem-allItem"]',
            )
        )
    )
    all_option.click()
    time.sleep(1)
    cal_year_month = driver.find_element(By.XPATH, '//span[@title="Cal. year / month"]')
    cal_year_month.click()
    time.sleep(1)
    all_option = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//span[@title="All"]'))
    )
    all_option.click()
    time.sleep(1)
    month_btn = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable(
            (By.CSS_SELECTOR, 'span[title="FEB 2023 (02.2023)"] bdi')
        )
    )
    ActionChains(driver).move_to_element(month_btn).click().perform()
    cal_year_month = driver.find_element(By.XPATH, '//span[@title="Cal. year / month"]')
    cal_year_month.click()
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//span[@id="__select0-labelText"]'))
    ).click()
    options = driver.find_elements(By.XPATH, '(//ul[@role="listbox"])[2]/child::*')
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//span[@id="__select0-labelText"]'))
    ).click()
    csv_path = os.path.join(project_dir, "_                   Vendor Rating Report.csv")
    for option in options:
        driver.find_element(By.CLASS_NAME, "sapMSltArrow").click()
        print(f"Downloading {option.text}")
        story_type = option.text
        option.click()
        time.sleep(2)
        driver.find_element(By.XPATH, '//span[contains(text(),"Measures")]').click()
        driver.find_element(By.XPATH, '//span[@title="More Actions"]').click()
        time.sleep(2)
        driver.find_element(By.XPATH, '//li[@title="Export"]').click()
        download_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//footer//button[1]"))
        )
        download_button.click()
        WebDriverWait(driver, 60).until(lambda x: os.path.exists(csv_path))
        df = pd.read_csv(csv_path)
        os.remove(csv_path)
        actual_cols = df.iloc[0].dropna().to_list()
        change_col = {df.columns[i]: actual_cols[i] for i in range(len(actual_cols))}
        df.rename(columns=change_col, inplace=True)
        df = df.iloc[1:]
        numeric_cols = df.select_dtypes(include="number").columns
        df = df[df["External group"] != "Not assigned"]
        for col in numeric_cols:
            if "%" in col:
                df[col] = df[col].map("{:.2f}".format)
        if story_type != "summary":
            df.drop(df.tail(1).index, inplace=True)
        df.to_excel(writer, sheet_name=story_type, index=False)


def sob_report_download(driver):
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located(
            (
                By.XPATH,
                "//div[@class='sapUiLocalBusyIndicatorAnimation sapUiLocalBusyIndicatorAnimStandard']",
            )
        )
    )
    cal_year_month = WebDriverWait(driver, 60).until(
        EC.element_to_be_clickable((By.XPATH, '//span[@title="Cal. year / month(1)"]'))
    )
    cal_year_month.click()
    all_option = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//span[@title="All"]'))
    )
    all_option.click()
    time.sleep(1)
    all_option = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//span[@title="All"]'))
    )
    all_option.click()
    time.sleep(1)
    search_month = driver.find_element(By.XPATH, '//input[@type="search"]')
    search_month.send_keys("FEB 2023")
    month_btn = driver.find_element(By.XPATH, '//bdi[normalize-space()="FEB 2023"]')
    month_btn.click()
    time.sleep(1)
    cal_year_month = driver.find_element(
        By.XPATH, '//span[@title="Cal. year / month(1)"]'
    )
    cal_year_month.click()
    time.sleep(1)

    for sob in SOB_TYPE:
        print(f"Downloading {sob} report")
        category = driver.find_element(By.XPATH, '//span[@title="CATEGORY(1)"]')
        category.click()
        time.sleep(1)
        all_option = driver.find_element(By.XPATH, '//span[@title="All"]')
        all_option.click()
        time.sleep(1)
        all_option = driver.find_element(By.XPATH, '//span[@title="All"]')
        all_option.click()

        driver.find_element(By.XPATH, '//input[@type="search"]').send_keys(sob)
        if sob == "POUCH":
            driver.find_element(
                By.XPATH,
                '//div[contains(@class,"sapEPMUIEffectiveFilterWidgetItem sapEPMUIEffectiveFilterWidgetItem-wordWrap")]',
            ).click()
        else:
            driver.find_element(
                By.XPATH, '//bdi[normalize-space()="{}"]'.format(sob)
            ).click()

        category = driver.find_element(By.XPATH, '//span[@title="CATEGORY(1)"]')
        category.click()
        time.sleep(2)
        driver.find_element(By.XPATH, '//span[contains(text(),"Measures")]').click()
        driver.find_element(By.XPATH, '//span[@title="More Actions"]').click()
        driver.find_element(By.XPATH, '//li[@title="Export"]').click()
        download_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//footer//button[1]"))
        )
        download_button.click()
        WebDriverWait(driver, 60).until(lambda x: os.path.exists(sob_download_path))
        df = pd.read_csv(sob_download_path, skipfooter=1)
        os.remove(sob_download_path)
        if sob != "ASEPTIC PACK":
            df = df.iloc[1:]
        sob_dict[sob] = df
    merged_df = pd.concat(sob_dict.values(), axis=0, ignore_index=True)
    actual_cols = merged_df.iloc[0].dropna().to_list()
    change_col = {merged_df.columns[i]: actual_cols[i] for i in range(len(actual_cols))}
    merged_df.rename(columns=change_col, inplace=True)
    merged_df = merged_df.iloc[1:]
    numeric_cols = merged_df.select_dtypes(include="number").columns
    merged_df[numeric_cols] = merged_df[numeric_cols].round(2)
    merged_df = merged_df[merged_df["Description"] != "Not assigned"]
    for col in merged_df.columns:
        if "%" in col:
            merged_df[col] *= 100
            merged_df[col] = merged_df[col].apply("{:.2f}".format)
    merged_df["RECEIPT QTY"] = merged_df["RECEIPT QTY"].apply("{:,.2f}".format)
    merged_df["RECEIPT VALUE"] = merged_df["RECEIPT VALUE"].apply("{:,.2f}".format)
    merged_df.to_excel(writer, index=False, sheet_name="SOB Report")


def format_workbook(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)
    header_font = Font(size=10, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(
        start_color="B6CFF2", end_color="B6CFF2", fill_type="solid"
    )
    cell_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    for sheet in wb:
        for column in sheet.columns:
            if column[0].value == "Description":
                for cell in column[1:]:
                    cell.alignment = Alignment(horizontal="left")
                    cell.font = Font(size=10)
                    cell.border = cell_border

            elif column[0].value in ["RECEIPT QTY", "RECEIPT VALUE"]:
                for cell in column[1:]:
                    cell.alignment = Alignment(horizontal="right")
                    cell.font = Font(size=10)
                    cell.border = cell_border
            else:
                for cell in column:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(size=10)
                    cell.border = cell_border
    for sheet in wb:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            if column_cells[0].value == "Vendor":
                vendor_col = column_cells[0].column_letter
                sheet.column_dimensions[vendor_col].width = max(
                    10, sheet.column_dimensions[vendor_col].width
                )
            else:
                sheet.column_dimensions[column_cells[0].column_letter].width = length
        if sheet.max_column > 7:
            for i in range(8, sheet.max_column + 1):
                col_letter = get_column_letter(i)
                sheet.column_dimensions[col_letter].width = 12
    for sheet in wb:
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = cell_border
        sheet.freeze_panes = "A2"
    wb.save(workbook_path)
    wb.close()


def run_main():
    global driver
    try:
        # Chrome preferences
        prefs = {
            "download.default_directory": download_path,
            "download.prompt_for_download": False,
        }

        user_agent = "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_2_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"
        Chrome_Options = webdriver.ChromeOptions()
        Chrome_Options.add_argument("headless")
        Chrome_Options.add_argument("--window-size=1920,1080")
        Chrome_Options.add_experimental_option("excludeSwitches", ["enable-logging"])
        Chrome_Options.add_experimental_option("excludeSwitches", ["enable-automation"])
        Chrome_Options.add_experimental_option("prefs", prefs)
        Chrome_Options.add_argument("--incognito")
        Chrome_Options.add_argument("user-agent=" + user_agent)

        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=Chrome_Options,
        )
        driver.implicitly_wait(2)
        driver.maximize_window()

        driver.get(os.getenv("WEBSITE_URL"))
        website_login(driver)
        print("Successfully logged into website...")

        WebDriverWait(driver, 60).until(
            EC.invisibility_of_element_located((By.ID, "__indicator0-busy-area"))
        )
        search1(driver)
        print("Searching for Vendor Rating Report...")

        WebDriverWait(driver, 60).until(
            EC.invisibility_of_element_located((By.ID, "__indicator0-busy-area"))
        )
        pass_story_credentials(driver)
        print("Filling up Story credentials...")

        WebDriverWait(driver, 60).until(
            EC.invisibility_of_element_located((By.ID, "__indicator0-busy-area"))
        )
        print("Downloading Vendor Rating Report")
        vendor_report_download(driver)
        print("Successfully Downloaded Vendor Rating Report \n")

        WebDriverWait(driver, 60).until(
            EC.invisibility_of_element_located((By.ID, "__indicator0-busy-area"))
        )
        search2(driver)
        print("Searching PM Vendor SOB Report New")

        WebDriverWait(driver, 60).until(
            EC.invisibility_of_element_located((By.ID, "__indicator0-busy-area"))
        )
        print("Downloading PM Vendor SOB Report New")
        sob_report_download(driver)
        print("Successfully downloaded PM Vendor SOB Report New")

        writer.close()
        format_workbook(filename)
        print("Sending mail...")
        send_mail(to, from_email, password, cc, subject, body, file_path=filename)
        end_time = time.time()
        print("Execution time: {} seconds".format(end_time - start_time))
    except Exception as e:
        driver.close()
        raise e
    finally:
        driver.quit()


if __name__ == "__main__":
    start_time = time.time()
    project_dir = os.getcwd()
    download_path = os.path.join(project_dir, "static")
    max_attempts = 3
    filename = "Vendor Rating Report (FEB 2023).xlsx"
    writer = pd.ExcelWriter(filename)
    sob_dict = {}
    SOB_TYPE = [
        "ASEPTIC PACK",
        "BOTTLE NEW",
        "BOTTLE OLD",
        "BOTTLE PET",
        "CANISTER",
        "CC BOX",
        "ENA",
        "HOLOGRAM",
        "LABEL",
        "MONOCARTON",
        "OTHERS",
        "PLASTIC CAP",
        "POUCH",
        "PP SEAL",
        "SHRINK SLEEVES",
    ]
    sob_download_path = os.path.join(project_dir, "PM Vendor Sob Report New.csv")
    from_email = os.getenv("SMTP_EMAIL_ADDRESS")
    # to = ['yoginderk@radico.co.in']
    to = ["mohaksharma@outlook.in", "mastwakrl@radico.co.in"]
    cc = []
    # cc = ['singhn@radico.co.in', 'bhattkc@radico.co.in', 'agarwalvk@radico.co.in']
    # cc = ['mastwalrk@radico.co.in', 'mohaksharma@outlook.in', 'singhn@radico.co.in']
    password = os.getenv("SMTP_PASSWORD")
    subject = "Vendor Rating Report (FEB 2023)"
    body = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <meta http-equiv="X-UA-Compatible" content="ie=edge">
    </head>
    <style>
      *{
        font-family:Helvetica, sans-serif;
        background-color:#f1f1f1;
      }
      body{
        padding:12px;
        box-shadow: rgba(0, 0, 0, 0.16) 0px 1px 4px, rgb(51, 51, 51) 0px 0px 0px 3px;
      }
    </style>
    <body>
        <h1>Vendor Rating Report (with SOB Report) - FEB 2023 </h1>
      <p> Please find the attached report with this mail. Also please note that <b>SOB Report</b> is the last sheet in attached workbook, so move accordingly.</p>
        <cite style='color:blue;'>(Do not reply to this as it is a automated message. Mail to mastwalrk@radico.co.in and mohaksharma@outlook.in for any further queries)</cite>
      <h3 style='font-style:italic;'>Thank You!</h3>
    </body>
    </html>
    """

    for attempt in range(max_attempts):
        try:
            print(f"Attempt number : {attempt + 1}")
            run_main()
        except Exception as e:
            print(e)
            if attempt == max_attempts - 1:
                print("All attempts exhausted")
            else:
                time.sleep(5)
        else:
            break
